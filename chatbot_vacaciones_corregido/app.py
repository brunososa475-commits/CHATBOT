import os
from datetime import datetime, date
from flask import Flask, render_template, request, redirect, url_for, session, jsonify
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = 'clave_secreta_para_sesiones_academicas'

# ==========================================
# CONFIGURACIÓN DE TU BASE DE DATOS REAL
# ==========================================
EXCEL_DB = r"C:\Users\milagros\OneDrive\Desktop\prueba y errr\chat bot\chatbot_vacaciones_corregido\data\base_de_datos_vacaciones.xlsx"

if not os.path.exists(EXCEL_DB):
    print(f"❌ ERROR CRÍTICO: No se encontró el archivo Excel en la ruta:\n{EXCEL_DB}")

# ==========================================
# FUNCIONES AUXILIARES DE EXCEL
# ==========================================
def _normalizar_a_date(val):
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    if isinstance(val, str):
        try:
            return datetime.strptime(val.strip()[:10], "%Y-%m-%d").date()
        except:
            return None
    return None

def buscar_empleado(legajo):
    try:
        wb = load_workbook(EXCEL_DB, data_only=True)
        ws = wb["empleados"] # Apunta a tu pestaña exacta
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            # row[0] es 'id_empleado', row[1] es 'nombre_apellido', row[2] es 'dias_disponibles'
            if row[0] is not None and str(row[0]).strip() == str(legajo).strip():
                return {
                    "legajo": str(row[0]), 
                    "nombre": row[1], 
                    "dias": int(row[2]) if row[2] is not None else 0, 
                    "email": row[3] if len(row) > 3 else "", 
                    "area": row[4] if len(row) > 4 else ""
                }
        return None
    except Exception as e:
        print(f"⚠️ Error al buscar empleado: {e}")
        return None

def actualizar_dias_empleado(legajo, dias_a_descontar):
    try:
        wb = load_workbook(EXCEL_DB)
        ws = wb["empleados"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value is not None and str(row[0].value).strip() == str(legajo).strip():
                dias_actuales = int(row[2].value) if row[2].value is not None else 0
                row[2].value = max(0, dias_actuales - dias_a_descontar)
                break
        wb.save(EXCEL_DB)
    except Exception as e:
        print(f"⚠️ No se pudieron descontar los días en el Excel: {e}")

def verificar_superposicion(legajo, inicio_nueva, fin_nueva):
    try:
        wb = load_workbook(EXCEL_DB, data_only=True)
        ws = wb["solicitud_de_vacaciones"] # Apunta a tu pestaña exacta
        
        in_nueva = datetime.strptime(inicio_nueva, "%Y-%m-%d").date()
        fn_nueva = datetime.strptime(fin_nueva, "%Y-%m-%d").date()

        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 5 or row[0] is None: 
                continue
            
            # row[1] es id_empleado, row[4] es estado_tramite
            if str(row[1]).strip() == str(legajo).strip() and str(row[4]).strip() in ["Pendiente", "Aprobado"]:
                in_existente = _normalizar_a_date(row[2])
                fn_existente = _normalizar_a_date(row[3])
                if in_existente and fn_existente:
                    if in_nueva <= fn_existente and fn_nueva >= in_existente:
                        return True
        return False
    except Exception as e:
        print(f"⚠️ Alerta en superposición: {e}")
        return False

def registrar_solicitud(legajo, inicio, fin):
    try:
        wb = load_workbook(EXCEL_DB)
        ws = wb["solicitud_de_vacaciones"]
        
        # 1. Buscamos cuál es el último ID y en qué fila está el verdadero final de los datos
        max_id = 0
        proxima_fila_libre = 2 # Arrancamos a buscar desde la fila 2
        
        for row_idx in range(2, ws.max_row + 2):
            celda_id = ws.cell(row=row_idx, column=1).value
            
            # Si la celda tiene un número o texto, seguimos buscando abajo
            if celda_id is not None and str(celda_id).strip() != "":
                if str(celda_id).isdigit():
                    max_id = max(max_id, int(celda_id))
                proxima_fila_libre = row_idx + 1
            else:
                # Si encontramos una celda vacía, encontramos el final real de los datos escritos
                proxima_fila_libre = row_idx
                break
        
        # 2. Generamos el nuevo ID correlativo
        nuevo_id = max_id + 1 if max_id > 0 else 105
        fecha_hoy = datetime.now().strftime("%Y-%m-%d")
        
        # 3. En lugar de usar append(), escribimos directamente en las celdas de la fila libre real
        ws.cell(row=proxima_fila_libre, column=1, value=nuevo_id)
        ws.cell(row=proxima_fila_libre, column=2, value=int(legajo))
        ws.cell(row=proxima_fila_libre, column=3, value=inicio)
        ws.cell(row=proxima_fila_libre, column=4, value=fin)
        ws.cell(row=proxima_fila_libre, column=5, value="Pendiente")
        ws.cell(row=proxima_fila_libre, column=6, value=fecha_hoy)
        
        wb.save(EXCEL_DB)
        print(f"✅ ¡Guardado real en la fila {proxima_fila_libre} con ID {nuevo_id}!")
    except Exception as e:
        print(f"❌ Error crítico al escribir en el Excel: {e}")
        raise e

# ==========================================
# PROCESAMIENTO LOGICA CHATBOT
# ==========================================
def procesar_mensaje_chatbot(mensaje):
    paso = session.get('paso', 'INICIO')
    historial = session.get('historial', [])
    mensaje = mensaje.strip()

    if paso == 'INICIO' and mensaje.lower() == 'hola':
        session['paso'] = 'ESPERANDO_LEGAJO'
        texto_bot = "¡Hola! Bienvenido al asistente de vacaciones. Por favor, ingresa tu número de ID de empleado para comenzar."
        historial.append({"emisor": "bot", "texto": texto_bot})
        session['historial'] = historial
        return

    if paso != 'INICIO':
        historial.append({"emisor": "user", "texto": mensaje})

    if paso == 'ESPERANDO_LEGAJO':
        empleado = buscar_empleado(mensaje)
        if empleado:
            session['empleado'] = empleado
            session['paso'] = 'ESPERANDO_FECHA_INICIO'
            texto_bot = f"Empleado encontrado: <b>{empleado['nombre']}</b>.<br>" \
                        f"Dispones de <b>{empleado['dias']} días disponibles</b>.<br><br>" \
                        f"Por favor, ingresá la <b>fecha de inicio</b> de tus vacaciones (Formato: AAAA-MM-DD):"
        else:
            texto_bot = "El ID de empleado no se encuentra registrado. Inténtalo de nuevo por favor."
        
    elif paso == 'ESPERANDO_FECHA_INICIO':
        try:
            fecha_in = datetime.strptime(mensaje, "%Y-%m-%d").date()
            if fecha_in < datetime.now().date():
                texto_bot = "La fecha de inicio no puede ser menor a la fecha actual. Ingresa una fecha válida (AAAA-MM-DD):"
            else:
                session['fecha_inicio'] = mensaje
                session['paso'] = 'ESPERANDO_FECHA_FIN'
                texto_bot = "Perfecto. Ahora ingresá la <b>fecha de fin</b> (Formato: AAAA-MM-DD):"
        except ValueError:
            texto_bot = "Formato inválido. Por favor usa el formato AAAA-MM-DD (Ej: 2026-07-20):"

    elif paso == 'ESPERANDO_FECHA_FIN':
        try:
            fecha_fn = datetime.strptime(mensaje, "%Y-%m-%d").date()
            fecha_in = datetime.strptime(session['fecha_inicio'], "%Y-%m-%d").date()
            
            if fecha_fn <= fecha_in:
                texto_bot = "La fecha de fin debe ser posterior a la fecha de inicio. Ingresa una nueva fecha de fin (AAAA-MM-DD):"
            else:
                dias_solicitados = (fecha_fn - fecha_in).days + 1
                dias_disponibles = session['empleado']['dias']
                
                if dias_solicitados > dias_disponibles:
                    texto_bot = f"Error: Has solicitado {dias_solicitados} días, pero solo dispones de {dias_disponibles}.<br>" \
                                f"Reiniciemos el periodo. Ingresa una nueva fecha de inicio (AAAA-MM-DD):"
                    session['paso'] = 'ESPERANDO_FECHA_INICIO'
                elif verificar_superposicion(session['empleado']['legajo'], session['fecha_inicio'], mensaje):
                    texto_bot = "Ya cuentas con una solicitud pendiente o aprobada que se superpone con estas fechas.<br>" \
                                "Por favor, ingresa una nueva fecha de inicio para evaluar (AAAA-MM-DD):"
                    session['paso'] = 'ESPERANDO_FECHA_INICIO'
                else:
                    session['fecha_fin'] = mensaje
                    session['dias_solicitados'] = dias_solicitados
                    session['paso'] = 'ESPERANDO_CONFIRMACION'
                    texto_bot = f"<b>Resumen de tu solicitud:</b><br>" \
                                f"• Periodo: {session['fecha_inicio']} al {session['fecha_fin']}<br>" \
                                f"• Total: {dias_solicitados} días seleccionados.<br><br>" \
                                f"¿Confirmas el pedido? Responde <b>SI</b> para enviar o <b>NO</b> para cancelar."
        except ValueError:
            texto_bot = "Formato inválido. Por favor usa el formato AAAA-MM-DD (Ej: 2026-09-15):"

    elif paso == 'ESPERANDO_CONFIRMACION':
        if mensaje.upper() in ['SI', 'SÍ', 'ACEPTAR', 'CONFIRMAR']:
            try:
                registrar_solicitud(session['empleado']['legajo'], session['fecha_inicio'], session['fecha_fin'])
                texto_bot = "¡Solicitud registrada con éxito! Ha quedado en estado 'Pendiente'. Ya podés cerrar el chat."
                session['paso'] = 'ESPERANDO_LEGAJO'
            except Exception:
                texto_bot = "❌ Ocurrió un problema al guardar. Asegurate de que el Excel no esté abierto."
        elif mensaje.upper() in ['NO', 'CANCELAR']:
            texto_bot = "Solicitud cancelada. Si deseas iniciar otra, ingresa tu ID de empleado nuevamente:"
            session['paso'] = 'ESPERANDO_LEGAJO'
        else:
            texto_bot = "Por favor responde únicamente <b>SI</b> o <b>NO</b>."

    historial.append({"emisor": "bot", "texto": texto_bot})
    session['historial'] = historial

# ==========================================
# RUTAS CONTROLADORAS (JSON)
# ==========================================
@app.route('/')
def index():
    session.clear()
    return render_template('index.html')

@app.route('/chat', methods=['POST'])
def chat():
    data = request.get_json() or {}
    mensaje_usuario = data.get('mensaje', '').strip()
    
    if 'historial' not in session:
        session['historial'] = []
        procesar_mensaje_chatbot("hola")
    elif mensaje_usuario:
        procesar_mensaje_chatbot(mensaje_usuario)
        
    historial = session.get('historial', [])
    ultimo_mensaje_bot = historial[-1]["texto"] if historial else "Error."
    
    return jsonify({
        "respuestas": [
            {"texto": ultimo_mensaje_bot}
        ]
    })

@app.route('/reiniciar', methods=['POST'])
def reiniciar():
    session.clear()
    return jsonify({"status": "reset exitoso"})

# ==========================================
# PANEL DE RECURSOS HUMANOS
# ==========================================
@app.route('/rrhh')
def rrhh_panel():
    try:
        wb = load_workbook(EXCEL_DB, data_only=True)
        ws = wb["solicitud_de_vacaciones"]
        solicitudes = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and str(row[4]).strip() == "Pendiente":
                solicitudes.append({
                    "id": int(row[0]), "legajo": row[1], "inicio": row[2], "fin": row[3], "estado": row[4], "fecha_sol": row[5]
                })
        return render_template('rrhh.html', solicitudes=solicitudes)
    except Exception as e:
        return f"Error en RRHH: {e}"

@app.route('/rrhh/aprobar/<int:id_solicitud>')
def aprobar_solicitud(id_solicitud):
    try:
        wb = load_workbook(EXCEL_DB)
        ws = wb["solicitud_de_vacaciones"]
        legajo, dias_a_descontar = None, 0
        se_modifico = False
        
        for row in ws.iter_rows(min_row=2):
            # Normalizamos el ID de la fila a texto para que compare bien siempre
            if row[0].value is not None and str(row[0].value).strip() == str(id_solicitud).strip():
                if str(row[4].value).strip() == "Pendiente":
                    row[4].value = "Aprobado"  # Modifica la columna 'estado_tramite'
                    legajo = row[1].value
                    in_date = _normalizar_a_date(row[2].value)
                    fn_date = _normalizar_a_date(row[3].value)
                    if in_date and fn_date:
                        dias_a_descontar = (fn_date - in_date).days + 1
                    se_modifico = True
                    break
                    
        if se_modifico and legajo:
            wb.save(EXCEL_DB)
            print(f"✅ Excel Guardado: Solicitud #{id_solicitud} pasada a Aprobado.")
            actualizar_dias_empleado(legajo, dias_a_descontar)
        else:
            wb.close()
            print(f"⚠️ No se encontró la solicitud #{id_solicitud} o ya no estaba Pendiente.")
    except Exception as e:
        print(f"❌ Error al aprobar en Excel: {e}")
        
    return redirect(url_for('rrhh_panel'))

@app.route('/rrhh/rechazar/<int:id_solicitud>')
def rechazar_solicitud(id_solicitud):
    try:
        wb = load_workbook(EXCEL_DB)
        ws = wb["solicitud_de_vacaciones"]
        se_modifico = False
        
        for row in ws.iter_rows(min_row=2):
            if row[0].value is not None and str(row[0].value).strip() == str(id_solicitud).strip():
                row[4].value = "Rechazado"  # Modifica la columna 'estado_tramite'
                se_modifico = True
                break
                
        if se_modifico:
            wb.save(EXCEL_DB)
            print(f"❌ Excel Guardado: Solicitud #{id_solicitud} pasada a Rechazado.")
        else:
            wb.close()
    except Exception as e:
        print(f"❌ Error al rechazar en Excel: {e}")
        
    return redirect(url_for('rrhh_panel'))

if __name__ == '__main__':
    app.run(debug=True)